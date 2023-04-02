import openpyxl


workbook = openpyxl.load_workbook("scores.xlsx")

worksheet = workbook["Список выпускников"]
worksheet2 = workbook["Школа и предметы"]


second_row_data = tuple(worksheet.rows)[1]

needed_data = []

last_row = worksheet.max_row
last_value = worksheet[f'A{last_row}'].value
# print(last_value)

needed_data = tuple(
    worksheet.iter_rows(min_row=2, min_col=6, max_row=last_row, values_only=True)
)
# needed_data = needed_data[0]

# Номер колонки с названиями школьных классов
class_column = 2

# Номер колонки с названиями предметов
subject_column = 3

# Создаем пустой список
matching_subjects = []

# Проходим циклом по всем строкам листа
for row in worksheet2.iter_rows(min_row=2, values_only=True):
    # Если значение в колонке "Класс" равно "9 Б класс"
    if row[class_column] == "9  Б класс":
        # Добавляем значение из колонки "Названия предметов" в список
        matching_subjects.append(row[subject_column])

matching_subjects = tuple(matching_subjects)

workbook.close()

# data = {"full_name_rus": worksheet["D2"].value, "full_name_kg": worksheet["E2"].value}
data = tuple(
    {
        'full_name_rus': worksheet[f'D{i}'].value,
        'full_name_kg': worksheet[f'E{i}'].value,
        'n_seria': worksheet[f'C{i}'].value
    } for i in range(2, last_row)
) 

quantity = 0

for student in data:
    scores = needed_data[quantity]
    for i in range(27):
        student.update(
            {
                i + 1: {
                    matching_subjects[i]: scores[i]
                }
            }
        )
    quantity += 1
